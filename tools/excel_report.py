"""
Camada 3 (Ferramentas) — Geração do relatório financeiro em Excel (.xlsx).

Recebe a lista de transações de um período e produz um arquivo Excel (.xlsx)
contendo duas abas:
  1. Resumo: totais de entradas, saídas, saldo e resumo agrupado por categoria.
  2. Transações: tabela detalhada de todos os lançamentos do período.

O arquivo é gravado no diretório temporário do SO (`tempfile.gettempdir()`),
portável tanto no local quanto na Vercel (/tmp).
"""
import os
import tempfile
from collections import defaultdict
from datetime import datetime
from zoneinfo import ZoneInfo


def _agrupar_por_categoria(transacoes: list) -> dict:
    grupos = defaultdict(lambda: {"entradas": 0.0, "saidas": 0.0})
    for t in transacoes:
        categoria = (t.get("categoria") or "").strip() or "Sem categoria"
        if t["status"] == "Entrada":
            grupos[categoria]["entradas"] += t["valor"]
        else:
            grupos[categoria]["saidas"] += abs(t["valor"])
    return grupos


def gerar_excel_relatorio(transacoes: list, data_inicio: str, data_fim: str, nome_usuario: str = "") -> str:
    """
    Gera o arquivo Excel (.xlsx) do relatório financeiro e retorna o caminho do arquivo criado.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    total_entradas = sum(t["valor"] for t in transacoes if t["status"] == "Entrada")

    total_saidas = sum(abs(t["valor"]) for t in transacoes if t["status"] == "Saída")
    saldo = total_entradas - total_saidas

    wb = Workbook()

    # --- Estilos ---
    font_titulo = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    font_subtitulo = Font(name="Arial", size=11, italic=True, color="555555")
    font_header = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    font_bold = Font(name="Arial", size=11, bold=True)
    font_normal = Font(name="Arial", size=11)

    fill_header_main = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    fill_header_sec = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    fill_accent = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9")
    )

    # --- ABA 1: RESUMO ---
    ws_resumo = wb.active
    ws_resumo.title = "Resumo"

    # Título do Relatório
    ws_resumo.merge_cells("A1:D1")
    ws_resumo["A1"] = "Relatório Financeiro - LekoAIFinance"
    ws_resumo["A1"].font = font_titulo
    ws_resumo["A1"].fill = fill_header_main
    ws_resumo["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_resumo.row_dimensions[1].height = 30

    if nome_usuario:
        ws_resumo["A2"] = f"Cliente: {nome_usuario}"
        ws_resumo["A2"].font = font_subtitulo
    ws_resumo["A3"] = f"Período: {data_inicio} a {data_fim}"
    ws_resumo["A3"].font = font_subtitulo

    # Tabela Resumo Geral
    ws_resumo["A5"] = "Resumo Geral"
    ws_resumo["A5"].font = font_bold

    resumo_dados = [
        ("Total de Entradas", total_entradas),
        ("Total de Saídas", total_saidas),
        ("Saldo do Período", saldo)
    ]

    for idx, (label, valor) in enumerate(resumo_dados, start=6):
        ws_resumo[f"A{idx}"] = label
        ws_resumo[f"B{idx}"] = valor
        ws_resumo[f"B{idx}"].number_format = 'R$ #,##0.00'
        ws_resumo[f"A{idx}"].font = font_bold if label.startswith("Saldo") else font_normal
        ws_resumo[f"B{idx}"].font = font_bold if label.startswith("Saldo") else font_normal
        if label.startswith("Saldo"):
            ws_resumo[f"A{idx}"].fill = fill_accent
            ws_resumo[f"B{idx}"].fill = fill_accent
        ws_resumo[f"A{idx}"].border = thin_border
        ws_resumo[f"B{idx}"].border = thin_border

    # Tabela Resumo por Categoria
    ws_resumo["A10"] = "Resumo por Categoria"
    ws_resumo["A10"].font = font_bold

    headers_cat = ["Categoria", "Entradas (R$)", "Saídas (R$)"]
    for col_num, h in enumerate(headers_cat, 1):
        cell = ws_resumo.cell(row=11, column=col_num)
        cell.value = h
        cell.font = font_header
        cell.fill = fill_header_sec
        cell.alignment = Alignment(horizontal="center" if col_num > 1 else "left")

    row_cat = 12
    for categoria, valores in sorted(_agrupar_por_categoria(transacoes).items()):
        c1 = ws_resumo.cell(row=row_cat, column=1, value=categoria)
        c2 = ws_resumo.cell(row=row_cat, column=2, value=valores['entradas'])
        c3 = ws_resumo.cell(row=row_cat, column=3, value=valores['saidas'])

        c2.number_format = 'R$ #,##0.00'
        c3.number_format = 'R$ #,##0.00'

        for c in (c1, c2, c3):
            c.font = font_normal
            c.border = thin_border
        row_cat += 1

    # --- ABA 2: TRANSAÇÕES ---
    ws_trans = wb.create_sheet(title="Transações")

    headers_trans = ["Data", "Tipo", "Valor (R$)", "Categoria", "Descrição"]
    for col_num, h in enumerate(headers_trans, 1):
        cell = ws_trans.cell(row=1, column=col_num)
        cell.value = h
        cell.font = font_header
        cell.fill = fill_header_main
        cell.alignment = Alignment(horizontal="center" if col_num in (1, 2) else "left")
    ws_trans.row_dimensions[1].height = 25

    for row_idx, t in enumerate(transacoes, start=2):
        c1 = ws_trans.cell(row=row_idx, column=1, value=t.get("data", ""))
        c2 = ws_trans.cell(row=row_idx, column=2, value=t.get("status", ""))
        c3 = ws_trans.cell(row=row_idx, column=3, value=abs(t.get("valor", 0.0)))
        c4 = ws_trans.cell(row=row_idx, column=4, value=t.get("categoria", ""))
        c5 = ws_trans.cell(row=row_idx, column=5, value=t.get("descricao", ""))

        c3.number_format = 'R$ #,##0.00'

        for c in (c1, c2, c3, c4, c5):
            c.font = font_normal
            c.border = thin_border

    # Ajuste automático de largura de colunas
    for ws in (ws_resumo, ws_trans):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or "")
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # Gravar arquivo temporário
    timestamp = datetime.now(ZoneInfo("America/Sao_Paulo")).strftime("%Y%m%d%H%M%S")
    caminho = os.path.join(tempfile.gettempdir(), f"relatorio_{timestamp}.xlsx")
    wb.save(caminho)
    return caminho
